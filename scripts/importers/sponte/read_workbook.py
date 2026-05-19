#!/usr/bin/env python3
import json
import sys
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


def normalize_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def main():
    if len(sys.argv) != 2:
        print(json.dumps({"error": "expected workbook path argument"}))
        sys.exit(1)

    workbook_path = Path(sys.argv[1])
    if not workbook_path.exists():
        print(json.dumps({"error": "workbook_not_found", "path": str(workbook_path)}))
        sys.exit(2)

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if not workbook.sheetnames:
        print(json.dumps({"error": "workbook_has_no_sheets", "path": str(workbook_path)}))
        sys.exit(3)

    worksheet = workbook[workbook.sheetnames[0]]
    rows_iter = worksheet.iter_rows(values_only=True)

    try:
        headers = next(rows_iter)
    except StopIteration:
      headers = tuple()

    header_list = []
    for index, header in enumerate(headers):
        if header is None or str(header).strip() == "":
            header_list.append(f"__empty_column_{index + 1}")
        else:
            header_list.append(str(header).strip())

    rows = []
    for row_index, values in enumerate(rows_iter, start=2):
        row = {}
        for column_name, value in zip(header_list, values):
            row[column_name] = normalize_value(value)
        rows.append({
            "source_row_number": row_index,
            "raw_payload": row,
        })

    print(json.dumps({
        "sheet_name": worksheet.title,
        "headers": header_list,
        "rows": rows,
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
